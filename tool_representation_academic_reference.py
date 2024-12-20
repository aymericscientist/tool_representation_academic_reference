# -*- coding: utf-8 -*-

import sys  # Import rapide pour gérer les sorties et erreurs critiques
from tkinter import Tk # Tk est utilisé pour gérer la fenêtre graphique sans afficher la fenêtre principale Tkinter
from tkinter.filedialog import askopenfilename  # Permet d'ouvrir une boîte de dialogue pour sélectionner un fichier


def load_txt_to_dataframe(file_path):
    """
    Charge les données d'un fichier TXT et les transforme en un DataFrame.
    """
    import pandas as pd # Bibliothèque utilisée pour manipuler et analyser des données sous forme de tableaux (DataFrame)
    from tqdm import tqdm # Permet d'afficher une barre de progression pour suivre l'avancement des boucles ou des traitements

    with open(file_path, 'r', encoding='utf-8') as file:
        lines = list(tqdm(file.readlines(), desc="Lecture des lignes du fichier TXT", unit="ligne"))

    # Construire un DataFrame avec une seule colonne : Source Reference
    data = {"Source Reference": [line.strip() for line in lines if line.strip()]}
    df = pd.DataFrame(data)
    return df


def calculate_similarity(source, target):
    """
    Calcule une similarité basée sur la correspondance des mots significatifs entre source et target.
    """
    import pandas as pd # Bibliothèque utilisée pour manipuler et analyser des données sous forme de tableaux (DataFrame)

    if pd.isna(source) or pd.isna(target):
        return 0

    source = str(source).lower()
    target = str(target).lower()

    target_words = [word for word in target.split() if len(word) > 3]

    matches = [word for word in target_words if word in source]
    return len(matches) / len(target_words) * 100 if target_words else 0


def resolve_doi(reference_queue, result_queue):
    """
    Résout les DOI pour une file de références en utilisant l'API CrossRef.
    """
    import time # Fournit des fonctions pour gérer les pauses et mesurer le temps écoulé
    import requests # Bibliothèque utilisée pour envoyer des requêtes HTTP et interagir avec des APIs ou des sites web

    url = "https://api.crossref.org/works"

    while not reference_queue.empty():
        reference = reference_queue.get()
        params = {"query": reference, "rows": 1}
        try:
            response = requests.get(url, params=params, timeout=10)
            response.raise_for_status()
            data = response.json()
            if (
                "message" in data
                and "items" in data["message"]
                and len(data["message"]["items"]) > 0
            ):
                item = data["message"]["items"][0]
                result_queue.put((reference, item.get("DOI", None)))
            else:
                result_queue.put((reference, None))
        except requests.exceptions.RequestException as e:
            print(f"Erreur lors de la résolution du DOI pour la référence '{reference}': {e}")
            result_queue.put((reference, None))

        time.sleep(1.5)  # Pause pour éviter de dépasser la limite de l'API


def fetch_metadata_from_doi(doi):
    """
    Récupère les métadonnées à partir d'un DOI en utilisant l'API CrossRef.
    """
    import requests # Bibliothèque utilisée pour envoyer des requêtes HTTP et interagir avec des APIs ou des sites web
    try:
        url = f"https://api.crossref.org/works/{doi}"
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        data = response.json()
        if "message" in data:
            item = data["message"]
            return {
                "Title": item.get("title", [None])[0],
                "Journal": item.get("container-title", [None])[0] if item.get("container-title", [None]) else None,
                "Authors": ", ".join(
                    [f"{author.get('given', '')} {author.get('family', '')}".strip() for author in item.get("author", [])]
                ),
                "pISSN_PRINT": next(
                    (issn.get("value") for issn in item.get("issn-type", []) if issn.get("type") == "print"),
                    None,
                ),
                "eISSN": next(
                    (issn.get("value") for issn in item.get("issn-type", []) if issn.get("type") == "electronic"),
                    None,
                ),
            }
        return {"Title": None, "Journal": None, "Authors": None, "pISSN_PRINT": None, "eISSN": None}
    except requests.exceptions.RequestException as e:
        print(f"Erreur lors de la récupération des métadonnées pour le DOI '{doi}': {e}")
        return {"Title": None, "Journal": None, "Authors": None, "pISSN_PRINT": None, "eISSN": None}


def process_references(df):
    """
    Résout les DOI et récupère les métadonnées pour chaque référence en mode asynchrone.
    """
    print("\nTraitement des références en cours...\n")
    from queue import Queue # Permet de gérer une file d'attente pour organiser et traiter les tâches de manière asynchrone
    from threading import Thread  # Permet d'exécuter plusieurs tâches en parallèle grâce au multi-threading
    from tqdm import tqdm # Utilisé pour afficher une barre de progression dans les boucles ou les opérations longues
    import pandas as pd # Bibliothèque utilisée pour manipuler et analyser des données sous forme de tableaux (DataFrame)

    reference_queue = Queue()
    result_queue = Queue()

    for reference in tqdm(df["Source Reference"], desc="Ajout des références à la file", unit="référence"):
        reference_queue.put(reference)

    threads = []
    for _ in range(5):  # Créer 5 threads pour paralléliser les requêtes
        thread = Thread(target=resolve_doi, args=(reference_queue, result_queue))
        thread.start()
        threads.append(thread)

    for thread in tqdm(threads, desc="Attente de la fin des threads", unit="thread"):
        thread.join()

    resolved_dois = {}
    while not result_queue.empty():
        reference, doi = result_queue.get()
        resolved_dois[reference] = doi

    metadata = []
    for reference in tqdm(df["Source Reference"], desc="Traitement des métadonnées", unit="référence"):
        doi = resolved_dois.get(reference)
        if doi:
            meta = fetch_metadata_from_doi(doi)
            meta["DOI"] = doi
            metadata.append(meta)
        else:
            metadata.append({"DOI": None, "Title": None, "Journal": None, "Authors": None, "pISSN_PRINT": None, "eISSN": None})

    metadata_df = pd.DataFrame(metadata)
    df = pd.concat([df, metadata_df], axis=1)

    print("Calcul des indices de similarité...")
    df["Similarity (Source vs Title)"] = df.apply(
        lambda row: calculate_similarity(row["Source Reference"], row["Title"]), axis=1
    )
    df["Similarity (Source vs Journal)"] = df.apply(
        lambda row: calculate_similarity(row["Source Reference"], row["Journal"]), axis=1
    )

    return df


def match_fnege_2022(df):
    """
    Ajoute une colonne 'Classement FNEGE 2022' en fonction des métadonnées.
    """
    import pandas as pd # Bibliothèque utilisée pour manipuler et analyser des données sous forme de tableaux (DataFrame)
    from tqdm import tqdm # Utilisé pour afficher une barre de progression dans les boucles ou les opérations longues

    url = "https://raw.githubusercontent.com/aymericscientist/tool_representation_academic_reference/0cfe2facb0100b05538a9b547e0fc3e1aa79a701/CLASSEMENT_FNEGE_2022_FORMATED.xlsx"

    print("Téléchargement du fichier FNEGE 2022 depuis GitHub...")
    try:
        fnege_data = pd.read_excel(url)
    except Exception as e:
        print(f"Erreur lors du téléchargement ou du chargement du fichier FNEGE 2022 : {e}")
        df["Classement FNEGE 2022"] = None
        return df

    fnege_data["TITLE"] = fnege_data["TITLE"].str.lower()
    fnege_data["pISSN_PRINT"] = fnege_data["pISSN_PRINT"].astype(str).str.lower()
    fnege_data["eISSN"] = fnege_data["eISSN"].astype(str).str.lower()

    df["Classement FNEGE 2022"] = None

    for index, row in tqdm(df.iterrows(), desc="Correspondance FNEGE", total=len(df)):
        pISSN = str(row["pISSN_PRINT"]).lower() if pd.notna(row["pISSN_PRINT"]) else None
        eISSN = str(row["eISSN"]).lower() if pd.notna(row["eISSN"]) else None
        journal = str(row["Journal"]).lower() if pd.notna(row["Journal"]) else None

        match = fnege_data[
            (fnege_data["pISSN_PRINT"] == pISSN) |
            (fnege_data["eISSN"] == eISSN) |
            (fnege_data["TITLE"] == journal)
        ]

        if not match.empty:
            df.at[index, "Classement FNEGE 2022"] = match.iloc[0]["Classement_2022"]

    return df


def export_to_excel(df, output_file):
    """
    Exporte le DataFrame avec une barre de progression au format Excel.
    """
    print("Export des données vers Excel...")
    from tqdm import tqdm # Utilisé pour afficher une barre de progression lors des traitements ou des boucles
    from openpyxl import Workbook # Permet de créer et manipuler des fichiers Excel au format .xlsx

    wb = Workbook()
    ws = wb.active
    ws.append(list(df.columns))
    for row in tqdm(df.values, desc="Écriture dans Excel", unit="ligne"):
        ws.append(row.tolist())

    wb.save(output_file)
    print(f"Les résultats ont été exportés dans : {output_file}")


def apply_conditional_formatting(output_file):
    """
    Applique une mise en forme conditionnelle aux colonnes d'indices de similarité dans le fichier Excel.
    """
    from openpyxl import load_workbook # Permet de charger un fichier Excel existant pour le lire ou le modifier
    from openpyxl.styles import PatternFill # Permet de définir des styles de remplissage (couleur) pour les cellules Excel

    wb = load_workbook(output_file)
    ws = wb.active

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for col_name in ["Similarity (Source vs Title)", "Similarity (Source vs Journal)"]:
        col_idx = [cell.column for cell in ws[1] if cell.value == col_name][0]
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is not None:
                if cell.value >= 80:
                    cell.fill = green_fill
                elif 60 <= cell.value < 80:
                    cell.fill = orange_fill
                elif cell.value < 60:
                    cell.fill = red_fill

    wb.save(output_file)
    print(f"Mise en forme conditionnelle appliquée dans : {output_file}")


def main():
    """
    Fonction principale.
    """
    Tk().withdraw()
    file_path = askopenfilename(title="Sélectionner le fichier TXT", filetypes=[("Text Files", "*.txt")])

    if not file_path:
        print("Aucun fichier sélectionné, le programme va se fermer.")
        return

    print(f"Lecture des données depuis le fichier : {file_path}")
    df = load_txt_to_dataframe(file_path)
    df = process_references(df)
    df = match_fnege_2022(df)
    output_file = file_path.replace(".txt", "_resolved_dois.xlsx")
    export_to_excel(df, output_file)
    apply_conditional_formatting(output_file)


if __name__ == "__main__":
    main()
    sys.exit(0)

    # Fin du script